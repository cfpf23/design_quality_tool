
import pandas as pd
from PIL import Image
from psd_tools import PSDImage
from datetime import datetime
import xlsxwriter
from openpyxl import load_workbook
import shutil, os


def now():
    """
    :return: str now
    """
    return datetime.now().strftime('%Y-%m-%d_%H-%M-%S')


def open_image(file: str) -> Image.Image:
    im = Image.open(file)
    return im



def save_data(df, path, templates):
    key = 'design_tool'
    date = datetime.today().strftime('%Y-%m-%d')
    df['date'] = date
    sheet_name = f'sheet_{key}'
    path_to_file = f"{path}/output_design_tool_{now()}.xlsx".replace("\\", "/")
    path_to_file = path_to_file.replace("-", "_")
    workbook = xlsxwriter.Workbook(path_to_file)
    # worksheet = workbook.add_worksheet(f'sheet_images_{date}')
    for k, v in templates.items():
        print(k, v)
        worksheet = workbook.add_worksheet(k)
        # template_image = open_image(v)
        #
        # image_width, image_height = template_image.size
        #
        # x_scale = ((945*image_width)/image_height) / image_width
        # y_scale = 945 / image_height
        #
        # worksheet.insert_image('B2', v,
        #                        {'x_scale': x_scale, 'y_scale': y_scale})
        worksheet.insert_image('B2', v)
    workbook.close()
    writer = pd.ExcelWriter(path_to_file, engine='openpyxl',mode='a')
    df.to_excel(writer, sheet_name=sheet_name, index=False)
    print(f'Processing data for {key} in function append_data')
    writer.save()

    return print(f'\nCorrectly appended data to {path} - {key}_report')


def check_presence_sheet(path_file, sheet_name):
    """
    retired function - do not erase
    """
    wb = load_workbook(path_file, read_only=True)
    if sheet_name in wb.sheetnames:
        return True
    else:
        return False


def add_templates_sheets(path, templates):
    """
    retired function - do not erase
    """
    workbook = load_workbook(path)
    for k, v in templates.items():
        sheet_name = f"template_{k}"
        if not check_presence_sheet(path, sheet_name):
            print(sheet_name)
            workbook.create_sheet(sheet_name)
            print(workbook.sheetnames)
            workbook.save(path)
            workbook.close()
    return


def append_data_new(df_from_folder, path_old_file, sheet_name,destination_to_report):
    """
    retired function - do not erase
    """
    wb = load_workbook(path_old_file, read_only=True)
    writer = pd.ExcelWriter(path_old_file, engine='xlsxwriter', options={'strings_to_urls': False})
    for sheet in wb.sheetnames:
        file = pd.read_excel(path_old_file, sheet_name=sheet)
        if sheet == sheet_name:
            file = file.append(df_from_folder, ignore_index=True)
            file.to_excel(writer, sheet_name=sheet, index=False, encoding='utf-8')
        else:
            file.to_excel(writer, sheet_name=sheet, index=False, encoding='utf-8')
    writer.save()
    # add_templates_sheets(path_old_file, templates)
    if destination_to_report:
        if '/'.join(path_old_file.split('/')[:-1]) != destination_to_report.replace('\\', '/'):
            shutil.move(path_old_file, destination_to_report)
            return
    else:
        return


# def append_right_sheet(df_from_folder, path_old_file, sheet_name, destination_to_report):
#     print(1)
#     print(sheet_name)
#     writer = pd.ExcelWriter(path_old_file, engine='xlsxwriter', options={'strings_to_urls': False})
#
#     file = pd.read_excel(path_old_file, sheet_name=sheet_name)
#     file = file.append(df_from_folder, ignore_index=True)
#     file.to_excel(writer, sheet_name=sheet_name, index=False, encoding='utf-8')
#     writer.save()
#     # add_templates(path_old_file)
#     if destination_to_report:
#         # destination = f'{destination_to_report}'
#         if '/'.join(path_old_file.split('/')[:-1]) != destination_to_report.replace('\\', '/'):
#             shutil.move(path_old_file, destination_to_report)
#             return
#     else:
#         return


def create_append_sheet(df_from_folder, path_old_file, sheet_name, destination_to_report):
    """
    retired function - do not erase
    """
    print(2)
    print(sheet_name)
    writer_old = pd.ExcelWriter(path_old_file, engine='openpyxl',mode='a')
    df_from_folder.to_excel(writer_old, sheet_name=sheet_name, index=False, encoding='utf-8')
    # writer_old.save()
    writer_old.close()
    # add_templates(path_old_file)
    # date = datetime.today().strftime('%Y-%m-%d')
    # add_templates_sheets(path_old_file, templates)
    if destination_to_report:
        # destination = f'{destination_to_report}'
        if '/'.join(path_old_file.split('/')[:-1]) != destination_to_report.replace('\\', '/'):
            shutil.move(path_old_file, destination_to_report)
            return
    else:
        return


def save_manager(df_from_folder, path_old_file, sheet_name, destination_to_report):

    if check_presence_sheet(path_old_file, sheet_name):
        append_data_new(df_from_folder, path_old_file, sheet_name, destination_to_report)
    else:
        create_append_sheet(df_from_folder, path_old_file, sheet_name, destination_to_report)


def save_and_append(df_from_folder, path_old_file, destination_to_report):
    date = datetime.today().strftime('%Y-%m-%d')
    df_from_folder['date'] = date
    sheet_name = f'sheet_images_{date}'

    save_manager(df_from_folder, path_old_file, sheet_name, destination_to_report)


def general_file_mapper(path):
    extensions = ["ase", "art", "bmp", "blp", "cd5", "cit", "cpt", "cr2", "cut", "dds",
                  "dib", "djvu", "egt", "exif", "gif", "gpl", "grf", "icns", "ico", "iff",
                  "jng", "jpeg", "jpg", "jfif", "jp2", "jps", "lbm", "max", "miff", "mng",
                  "msp", "nitf", "ota", "pbm", "pc1", "pc2", "pc3", "pcf", "pcx", "pdn",
                  "pgm", "PI1", "PI2", "PI3", "pict", "pct", "pnm", "pns", "ppm", "psb",
                  "psd", "pdd", "psp", "px", "pxm", "pxr", "qfx", "raw", "rle", "sct",
                  "sgi", "rgb", "int", "bw", "tga", "tiff", "tif", "vtf", "xbm", "xcf",
                  "xpm", "3dv", "amf", "ai", "awg", "cgm", "cdr", "cmx", "dxf", "e2d",
                  "egt", "eps", "fs", "gbr", "odg", "svg", "stl", "vrml", "x3d", "sxd",
                  "v2d", "vnd", "wmf", "emf", "art", "xar", "png", "webp", "jxr", "hdp",
                  "wdp", "cur", "ecw", "iff", "lbm", "liff", "nrrd", "pam", "pcx", "pgf",
                  "sgi", "rgb", "rgba", "bw", "int", "inta", "sid", "ras", "sun", "tga"]
    result = []
    for (root, dirs, files) in os.walk(path, topdown=True):
        f = [root + "\\" + file for file in files]
        if len(f) > 0:
            for image in f:
                result.append({'file_origin': '\\'.join(root.split('\\')[5:]),
                               'file_address': image,
                               'file_type': image.split('.')[-1],
                               'is_image': 1 if image.split('.')[-1] in extensions else 0})
    return result


def byte_organizer(number: int) -> str:
    try:
        if len(str(number)) <= 6:
            return f"{str(round(number/1000, 3))} KB"
        elif (len(str(number)) > 6) & (len(str(number)) <= 9):
            return f"{str(round(number/1000000, 3))} MB"
        elif (len(str(number)) > 9) & (len(str(number)) <= 12):
            return f"{str(round(number/1000000000, 3))} GB"
    except Exception as e:
        print(e)
        return str(e)


def file_size_getter(file: str):
    try:
        file = file.replace("\\", "/")
        if os.path.getsize(file) > 150000:
            size = os.path.getsize(file) - os.path.getsize('/'.join(file.split('/')[:-1]))
            return size
        else:
            return os.path.getsize(file)
    except Exception as e:
        return str(e)


def get_image_size(row_file):
    memory_usage = file_size_getter(row_file)
    memory_usage = byte_organizer(memory_usage)
    width = '-'
    height = '-'
    try:
        if row_file.split('.')[-1] in ['jpg', 'png']:
            im = Image.open(row_file)
            width, height = im.size
            width = str(width)
            height = str(height)
            return f'{width},{height},{memory_usage}'
        elif row_file.split('.')[-1] == 'psd':
            img_psd = PSDImage.load(row_file)
            width = img_psd.header.width
            height = img_psd.header.height
            return f'{width},{height},{memory_usage}'
    except:
        print(row_file)
        return f'{width},{height},{memory_usage}'


def categorization(string):
    categorization = {
        r'A+\Escada\Poprawne pliki':'A+',
        r'A+\Escada\Niepoprawne pliki'  :'A+',
        "Banery":"Banners",
        r'Brandstrore\PSD':'Brandstore',
        r'Brandstrore\PNG':'Brandstore',
        'Basic Content':'Basic Content',
        'Templatki z formatami':'Template with format'
    }
    try:
        new_string = categorization[string]
        return new_string
    except:
        return string


def dataframe_generation(path):
    df = pd.DataFrame(general_file_mapper(path))
    df[['width', 'height', 'memory_usage_bytes']] = df['file_address'].apply(get_image_size).str.split(',', expand=True)
    df['categorization'] = df['file_origin'].apply(categorization)
    df.categorization.fillna('other', inplace=True)
    return df


def save_file(df, path):
    """
    retired function - do not erase
    """

    date = datetime.today().strftime('%Y-%m-%d')
    destination = f'{path}/report_images_{date}.xlsx'
    df['date'] = date
    if os.path.exists(destination):
        writer = pd.ExcelWriter(f"{destination}", engine='xlsxwriter', options={'strings_to_urls': False})
    else:
        # workbook = xlsxwriter.Workbook(destination)
        # worksheet = workbook.add_worksheet(f'sheet_images_{date}')
        # for k, v in templates.items():
        #     print(k, v)
        #     worksheet = workbook.add_worksheet(k)
        #     worksheet.insert_image('B2', v)
        # workbook.close()
        writer = pd.ExcelWriter(f"{destination}", engine='xlsxwriter', options={'strings_to_urls': False})
    print(f'Processing data for images in function append_data')
    file = pd.read_excel(destination, sheet_name=f'sheet_images_{date}')
    file = file.append(df, ignore_index=True)
    file.to_excel(writer, sheet_name=f'sheet_images_{date}', index=False, encoding='utf-8')
    # file.to_excel(path, sheet_name=sheet, index=False)
    writer.save()

    return print(f'\nCorrectly appended data to {destination} - images_report')


def file_reader(file):
    """
    :param file: read the file path. Must csv or xlsx.
    :return: list containinf the dataframe and the date, if not able to read, it will return the file name
    """
    if file.endswith('.csv'):
        df = pd.read_csv(file)
        return df
    elif file.endswith('.xlsx'):
        df = pd.read_excel(file)
        return df
    else:
        print(f'not able to read {file}')
        return False

